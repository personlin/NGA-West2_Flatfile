#!/usr/bin/env python3
"""Catalog workbook formulas relevant to the NGA-West2 GMPE workbook."""
from __future__ import annotations

import argparse
import csv
import hashlib
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("data/NGAW2_GMPE_Spreadsheets_v5.7_041415_ProtectedLocked.xlsm")
OUT = Path("gmpe/reference/workbook_formula_catalog.csv")


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--out", type=Path, default=OUT)
    args = parser.parse_args()

    workbook_hash = sha256(args.workbook)
    wb = load_workbook(args.workbook, data_only=False, read_only=False, keep_vba=False)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    rows.append(
                        {
                            "source_workbook": args.workbook.name,
                            "source_workbook_sha256": workbook_hash,
                            "sheet": sheet,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                        }
                    )

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["source_workbook", "source_workbook_sha256", "sheet", "cell", "formula"],
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {len(rows)} formulas to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
