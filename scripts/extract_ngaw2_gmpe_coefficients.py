#!/usr/bin/env python3
"""Extract NGA-West2 GMPE workbook coefficient sheets to CSV files."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK = Path("data/NGAW2_GMPE_Spreadsheets_v5.7_041415_ProtectedLocked.xlsm")
OUT_DIR = Path("gmpe/coefficients")

MODEL_SHEETS = {
    "ASK14": "ASK14_Coeffs",
    "BSSA14": "BSSA14_Coeffs",
    "CB14": "CB14_Coeffs",
    "CY14": "CY14_Coeffs",
    "I14": "I14_Coeffs",
}

CSV_COLUMNS = [
    "model",
    "period_s",
    "period_key",
    "imt",
    "row_index",
    "coefficient",
    "value",
    "cached_value",
    "formula",
    "source_sheet",
    "source_cell",
    "source_workbook",
]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def period_key(period: Any) -> str:
    value = float(period)
    if value == 0:
        return "pga"
    if value == -1:
        return "pgv"
    sign = "m" if value < 0 else "p"
    return f"{sign}{abs(value):.3f}".replace(".", "p")


def imt_for_period(period: Any) -> str:
    value = float(period)
    if value == 0:
        return "PGA"
    if value == -1:
        return "PGV"
    return "SA"


def cell_value(cell: Any) -> Any:
    value = cell.value
    if value == "":
        return None
    return value


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def duplicate_stable_names(values: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for value in values:
        name = str(value).strip() if value not in (None, "") else "unnamed"
        seen[name] = seen.get(name, 0) + 1
        out.append(name if seen[name] == 1 else f"{name}_{seen[name]}")
    return out


def used_bounds(ws: Any) -> tuple[int, int, int, int]:
    rows: list[int] = []
    cols: list[int] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell_value(cell) is not None:
                rows.append(cell.row)
                cols.append(cell.column)
    if not rows:
        return (1, 1, 1, 1)
    return (min(rows), max(rows), min(cols), max(cols))


def numeric_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def extract_model(
    model: str,
    sheet: str,
    formulas_wb: Any,
    values_wb: Any,
    workbook_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    ws_formula = formulas_wb[sheet]
    ws_value = values_wb[sheet]
    min_row, max_row, min_col, max_col = used_bounds(ws_formula)

    header_row = 3
    first_data_row = 5
    headers = duplicate_stable_names(
        [ws_formula.cell(header_row, col).value for col in range(min_col, max_col + 1)]
    )

    records: list[dict[str, Any]] = []
    formula_count = 0
    non_empty_count = 0
    periods: list[float] = []

    for row in range(first_data_row, max_row + 1):
        period = numeric_or_none(ws_value.cell(row, min_col).value)
        if period is None:
            continue
        periods.append(period)
        key = period_key(period)
        imt = imt_for_period(period)
        for offset, col in enumerate(range(min_col + 1, max_col + 1), start=1):
            src = ws_formula.cell(row, col)
            cached = ws_value.cell(row, col)
            raw_value = cell_value(src)
            cached_value = cell_value(cached)
            formula = raw_value if is_formula(raw_value) else None
            value = None if formula else raw_value
            if raw_value is not None or cached_value is not None:
                non_empty_count += 1
            if formula is not None:
                formula_count += 1
            if raw_value is None and cached_value is None:
                continue
            records.append(
                {
                    "model": model,
                    "period_s": period,
                    "period_key": key,
                    "imt": imt,
                    "row_index": row,
                    "coefficient": headers[offset],
                    "value": value,
                    "cached_value": cached_value,
                    "formula": formula,
                    "source_sheet": sheet,
                    "source_cell": f"{get_column_letter(col)}{row}",
                    "source_workbook": workbook_path.name,
                }
            )

    manifest = {
        "model": model,
        "source_sheet": sheet,
        "used_range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
        "header_row": header_row,
        "first_data_row": first_data_row,
        "period_count": len(set(periods)),
        "row_count": len(records),
        "formula_cell_count": formula_count,
        "non_empty_data_cell_count": non_empty_count,
    }
    return records, manifest


def write_csv(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(records)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    workbook = args.workbook
    out_dir = args.out_dir
    if not workbook.exists():
        raise FileNotFoundError(workbook)

    out_dir.mkdir(parents=True, exist_ok=True)
    formulas_wb = load_workbook(workbook, data_only=False, read_only=False, keep_vba=False)
    values_wb = load_workbook(workbook, data_only=True, read_only=False, keep_vba=False)

    workbook_hash = sha256(workbook)
    sheets: list[dict[str, Any]] = []
    for model, sheet in MODEL_SHEETS.items():
        records, sheet_manifest = extract_model(model, sheet, formulas_wb, values_wb, workbook)
        write_csv(out_dir / f"{model.lower()}.csv", records)
        sheets.append(sheet_manifest)

    manifest = {
        "source_workbook": workbook.name,
        "source_workbook_sha256": workbook_hash,
        "extracted_at_utc": datetime.now(timezone.utc).isoformat(),
        "csv_schema": CSV_COLUMNS,
        "sheets": sheets,
    }
    (out_dir / "coefficient_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote coefficient CSVs and manifest to {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
